"""
extrair_brindes_playwright.py
- Usa Playwright para abrir Outlook Web (reusa sessão salvo em outlook_state.json).
- Procura e-mails cujo assunto contenha "Lista de brindes" (do dia anterior preferencialmente).
- Extrai a primeira tabela HTML presente no corpo do e-mail.
- Salva os dados em Brindes.xlsx (cria se não existir).
"""

from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError
from bs4 import BeautifulSoup
import pandas as pd
import datetime, os, time
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

OUTLOOK_URL = "https://outlook.office.com/mail/"
STORAGE_STATE = "outlook_state.json"
ARQUIVO_PLANILHA = "Brindes.xlsx"
ASSUNTO_CHAVE = "Lista de brindes"  # texto usado no assunto (case-insensitive)

def ensure_logged_in(playwright):
    """
    Abre browser para login manual se storage state não existir.
    Salva storage state em outlook_state.json após o login manual.
    """
    browser = playwright.chromium.launch(headless=False)
    context = browser.new_context()
    page = context.new_page()
    page.goto(OUTLOOK_URL, timeout=60000)
    print(">>> Janela do navegador aberta. Faça login manualmente no Outlook (SSO/MFA se houver).")
    print(">>> Depois de logar e ver sua caixa de entrada, volte aqui e pressione ENTER para salvar a sessão.")
    input("Pressione ENTER depois de concluir o login no navegador... ")
    # opcional: esperar um pouco para garantir carregamento
    time.sleep(2)
    context.storage_state(path=STORAGE_STATE)
    print(f"✅ Sessão salva em {STORAGE_STATE}")
    context.close()
    browser.close()

def open_with_saved_session(playwright):
    browser = playwright.chromium.launch(headless=False)
    context = browser.new_context(storage_state=STORAGE_STATE)
    page = context.new_page()
    page.goto(OUTLOOK_URL, timeout=60000)
    # aguarda leitura estar disponível
    page.wait_for_load_state('load', timeout=60000)
    return browser, context, page

def find_latest_brindes_email(page):
    """
    Tenta localizar o e-mail mais recente cujo assunto contenha ASSUNTO_CHAVE.
    Estratégia:
      - Tenta usar a caixa de busca superior para buscar por ASSUNTO_CHAVE
      - Abre o primeiro resultado (se houver)
      - Retorna True se um e-mail foi aberto
    """
    queries = [ASSUNTO_CHAVE, f'subject:"{ASSUNTO_CHAVE}"']
    search_selectors = [
        # tenta alguns seletores possíveis para a caixa de busca
        'input[placeholder="Pesquisar"]',
        'input[aria-label="Pesquisar"]',
        'input[aria-label="Search"]',
        'input[type="search"]'
    ]

    # tentar pesquisar
    for q in queries:
        for sel in search_selectors:
            try:
                search = page.locator(sel)
                if search.count() == 0:
                    continue
                search.click()
                search.fill(q)
                search.press("Enter")
                # aguardar resultados carregarem
                time.sleep(2)
                # primeiro item da lista de mensagens
                # Outlook usa elementos com role="listitem" para cada e-mail
                # vamos pegar o primeiro que contenha o texto do assunto
                items = page.locator('div[role="listitem"]')
                # filtra por subject text
                candidate = None
                count = items.count()
                for i in range(count):
                    text = items.nth(i).inner_text(timeout=2000)
                    if ASSUNTO_CHAVE.lower() in text.lower():
                        candidate = items.nth(i)
                        break
                if candidate:
                    candidate.click()
                    # aguarda painel de leitura carregar
                    time.sleep(1)
                    return True
            except PlaywrightTimeoutError:
                continue
            except Exception:
                continue
    # fallback: varrer listagem principal da inbox
    try:
        inbox_items = page.locator('div[role="listitem"]')
        for i in range(inbox_items.count()):
            txt = inbox_items.nth(i).inner_text(timeout=2000)
            if ASSUNTO_CHAVE.lower() in txt.lower():
                inbox_items.nth(i).click()
                time.sleep(1)
                return True
    except Exception:
        pass

    return False

def extract_html_from_read_pane(page):
    """
    Tenta localizar o conteúdo HTML do e-mail (painel de leitura).
    Retorna HTML string ou None.
    """
    selectors_try = [
        'div[role="document"]',               # comum
        'div[aria-label="Mensagem"]',         # pt-br possibilidade
        'div[aria-label="Message"]',          # en
        'div[aria-label="Message body"]',
        'div[aria-label="Message pane"]',
        'div[data-test-id="message-body"]'
    ]
    for sel in selectors_try:
        try:
            loc = page.locator(sel)
            if loc.count() > 0:
                html = loc.first.inner_html(timeout=3000)
                if "<table" in html.lower():
                    return html
        except Exception:
            continue
    # último recurso: pegar HTML do painel direito inteiro
    try:
        right_pane = page.locator('div[role="main"]')
        if right_pane.count() > 0:
            html = right_pane.first.inner_html(timeout=3000)
            if "<table" in html.lower():
                return html
    except Exception:
        pass
    return None

def parse_first_table(html):
    """
    Usa pandas.read_html via BeautifulSoup-subset para extrair a primeira tabela.
    Retorna dataframe ou None.
    """
    try:
        # isolar a primeira tabela com BeautifulSoup
        soup = BeautifulSoup(html, "html.parser")
        table = soup.find("table")
        if table is None:
            return None
        table_html = str(table)
        df_list = pd.read_html(table_html)
        if len(df_list) > 0:
            df = df_list[0]
            return df
    except Exception as e:
        print("Erro ao parsear tabela:", e)
    return None

def save_dataframe_to_excel(df, remetente, assunto, data_referencia):
    """
    Salva (ou anexa) DataFrame em ARQUIVO_PLANILHA.
    Adiciona colunas de metadados: DataRecebimento, Remetente, Assunto, DataReferencia
    """
    # normalizar df (colunas com nomes simples)
    df = df.copy()
    df.insert(0, "DataRecebimento", datetime.datetime.now().strftime("%d/%m/%Y %H:%M"))
    df.insert(1, "Remetente", remetente)
    df.insert(2, "Assunto", assunto)
    df.insert(3, "DataReferencia", data_referencia)

    if os.path.exists(ARQUIVO_PLANILHA):
        # abrir e anexar
        with pd.ExcelWriter(ARQUIVO_PLANILHA, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
            # escreve em nova aba com timestamp
            sheet_name = "Brindes"
            # se existir, vamos carregar e concatenar manualmente
            try:
                existing = pd.read_excel(ARQUIVO_PLANILHA, sheet_name=sheet_name)
                combined = pd.concat([existing, df], ignore_index=True)
                combined.to_excel(writer, sheet_name=sheet_name, index=False)
            except Exception:
                # se falhar em ler, apenas escreve num sheet novo com timestamp
                ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                df.to_excel(writer, sheet_name=f"Brindes_{ts}", index=False)
    else:
        # cria novo arquivo
        df.to_excel(ARQUIVO_PLANILHA, sheet_name="Brindes", index=False)
        # abrir e formatar header com openpyxl
        from openpyxl import load_workbook
        wb = load_workbook(ARQUIVO_PLANILHA)
        ws = wb["Brindes"]
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CFE2F3", end_color="CFE2F3", fill_type="solid")
        wb.save(ARQUIVO_PLANILHA)

    print(f"✅ Dados salvos em {ARQUIVO_PLANILHA}")

def main():
    with sync_playwright() as pw:
        # se não houver sessão salva, pedir para logar manualmente e salvar
        if not os.path.exists(STORAGE_STATE):
            print("Não foi encontrada uma sessão salva (outlook_state.json). Vamos abrir o navegador para você logar.")
            ensure_logged_in(pw)

        # abrir com sessão salva
        browser, context, page = open_with_saved_session(pw)
        try:
            opened = find_latest_brindes_email(page)
            if not opened:
                print("⚠️ Não encontrei e-mail com assunto contendo:", ASSUNTO_CHAVE)
                context.close()
                browser.close()
                return

            # pegar metadados do e-mail visíveis na UI (remetente e assunto)
            remetente = ""
            assunto = ""
            try:
                # tenta pegar elementos que contenham remetente/assunto no painel
                subj_loc = page.locator('h1').first
                assunto = subj_loc.inner_text(timeout=2000)
            except Exception:
                try:
                    assunto = page.locator('div[role="heading"]').first.inner_text(timeout=2000)
                except Exception:
                    assunto = ASSUNTO_CHAVE

            try:
                remetente = page.locator('div[role="button"][title]').first.get_attribute("title") or ""
            except Exception:
                # fallback: tentar texto do cabeçalho
                try:
                    header = page.locator('div[aria-label="Reading pane"]').first.inner_text(timeout=2000)
                    remetente = header.splitlines()[0] if header else ""
                except Exception:
                    remetente = ""

            # extrair html do corpo
            html = extract_html_from_read_pane(page)
            if not html:
                print("⚠️ Não consegui extrair HTML do corpo do e-mail.")
                context.close()
                browser.close()
                return

            df = parse_first_table(html)
            if df is None:
                print("⚠️ Nenhuma tabela encontrada no corpo do e-mail.")
                context.close()
                browser.close()
                return

            # opcional: detectar data de referência no texto (ex: 'referente ao dia 07/11/2025')
            data_referencia = ""
            try:
                text_all = BeautifulSoup(html, "html.parser").get_text(" ", strip=True)
                import re
                m = re.search(r"(\d{2}/\d{2}/\d{4})", text_all)
                if m:
                    data_referencia = m.group(1)
            except Exception:
                data_referencia = ""

            save_dataframe_to_excel(df, remetente, assunto, data_referencia)
        finally:
            context.close()
            browser.close()

if __name__ == "__main__":
    main()
